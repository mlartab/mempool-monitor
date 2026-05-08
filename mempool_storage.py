"""
mempool_storage.py
──────────────────
Persistent storage for mempool transactions.
Writes to SQLite immediately (never lost), exports to CSV / JSON / Excel on schedule.

Install extras:  pip install openpyxl pandas
"""

import csv
import json
import os
import sqlite3
import threading
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ── Paths ───────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent / "mempool_data"
DB_PATH     = BASE_DIR / "mempool.db"
EXPORT_DIR  = BASE_DIR / "exports"

BASE_DIR.mkdir(exist_ok=True)
EXPORT_DIR.mkdir(exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════
# DATABASE
# ═══════════════════════════════════════════════════════════════════════
class MempoolDB:
    """Thread-safe SQLite wrapper. Every TX is committed immediately."""

    def __init__(self, db_path: Path = DB_PATH):
        self.db_path = db_path
        self._local  = threading.local()
        self._init_schema()

    def _conn(self) -> sqlite3.Connection:
        if not hasattr(self._local, "conn") or self._local.conn is None:
            self._local.conn = sqlite3.connect(
                str(self.db_path),
                check_same_thread=False,
                timeout=10,
            )
            self._local.conn.row_factory = sqlite3.Row
            self._local.conn.execute("PRAGMA journal_mode=WAL")   # faster writes
            self._local.conn.execute("PRAGMA synchronous=NORMAL") # safe + fast
        return self._local.conn

    def _init_schema(self):
        c = self._conn()
        c.executescript("""
            CREATE TABLE IF NOT EXISTS transactions (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                hash        TEXT    UNIQUE NOT NULL,
                from_addr   TEXT,
                to_addr     TEXT,
                value_eth   REAL    DEFAULT 0,
                gas_gwei    REAL    DEFAULT 0,
                func_sig    TEXT,
                func_name   TEXT,
                risk        TEXT    DEFAULT 'low',
                flags       TEXT,          -- JSON array stored as string
                seen_at     TEXT    NOT NULL,
                block_number INTEGER
            );

            CREATE TABLE IF NOT EXISTS alerts (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                tx_hash     TEXT    NOT NULL,
                risk        TEXT,
                flags       TEXT,
                value_eth   REAL,
                from_addr   TEXT,
                alerted_at  TEXT    NOT NULL
            );

            CREATE TABLE IF NOT EXISTS stats_hourly (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                hour            TEXT    UNIQUE NOT NULL,  -- YYYY-MM-DD HH
                total_tx        INTEGER DEFAULT 0,
                high_value_tx   INTEGER DEFAULT 0,
                drain_sig_tx    INTEGER DEFAULT 0,
                critical_tx     INTEGER DEFAULT 0,
                total_eth       REAL    DEFAULT 0
            );

            CREATE INDEX IF NOT EXISTS idx_tx_risk    ON transactions(risk);
            CREATE INDEX IF NOT EXISTS idx_tx_seen    ON transactions(seen_at);
            CREATE INDEX IF NOT EXISTS idx_tx_from    ON transactions(from_addr);
            CREATE INDEX IF NOT EXISTS idx_tx_func    ON transactions(func_sig);
        """)
        c.commit()

    # ── Write ────────────────────────────────────────────────────────
    def save_tx(self, tx: dict):
        """Insert a classified transaction. Ignores duplicates."""
        c = self._conn()
        try:
            c.execute("""
                INSERT OR IGNORE INTO transactions
                    (hash, from_addr, to_addr, value_eth, gas_gwei,
                     func_sig, func_name, risk, flags, seen_at)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (
                tx.get("hash", ""),
                tx.get("from", ""),
                tx.get("to", ""),
                float(tx.get("value_eth", 0)),
                float(tx.get("gas_gwei", 0)),
                tx.get("func_sig", ""),
                tx.get("func_name", "unknown"),
                tx.get("risk", "low"),
                json.dumps(tx.get("flags", [])),
                tx.get("timestamp", datetime.utcnow().isoformat()),
            ))
            c.commit()
            self._update_hourly_stats(tx)
        except sqlite3.Error as e:
            print(f"[DB] Write error: {e}")

    def save_alert(self, tx: dict):
        """Log an alert separately for quick querying."""
        c = self._conn()
        try:
            c.execute("""
                INSERT INTO alerts
                    (tx_hash, risk, flags, value_eth, from_addr, alerted_at)
                VALUES (?,?,?,?,?,?)
            """, (
                tx.get("hash", ""),
                tx.get("risk", ""),
                json.dumps(tx.get("flags", [])),
                float(tx.get("value_eth", 0)),
                tx.get("from", ""),
                datetime.utcnow().isoformat(),
            ))
            c.commit()
        except sqlite3.Error as e:
            print(f"[DB] Alert write error: {e}")

    def _update_hourly_stats(self, tx: dict):
        hour = datetime.utcnow().strftime("%Y-%m-%d %H")
        c    = self._conn()
        c.execute("""
            INSERT INTO stats_hourly (hour, total_tx, total_eth) VALUES (?,1,?)
            ON CONFLICT(hour) DO UPDATE SET
                total_tx      = total_tx + 1,
                total_eth     = total_eth + excluded.total_eth,
                high_value_tx = high_value_tx + CASE WHEN ? THEN 1 ELSE 0 END,
                drain_sig_tx  = drain_sig_tx  + CASE WHEN ? THEN 1 ELSE 0 END,
                critical_tx   = critical_tx   + CASE WHEN ? THEN 1 ELSE 0 END
        """, (
            hour,
            float(tx.get("value_eth", 0)),
            "HIGH_VALUE"      in tx.get("flags", []),
            "TOKEN_DRAIN_SIG" in tx.get("flags", []),
            tx.get("risk") == "critical",
        ))
        c.commit()

    # ── Query ────────────────────────────────────────────────────────
    def get_recent(self, limit: int = 100) -> list[dict]:
        rows = self._conn().execute(
            "SELECT * FROM transactions ORDER BY id DESC LIMIT ?", (limit,)
        ).fetchall()
        return [dict(r) for r in rows]

    def get_by_risk(self, risk: str, limit: int = 100) -> list[dict]:
        rows = self._conn().execute(
            "SELECT * FROM transactions WHERE risk=? ORDER BY id DESC LIMIT ?",
            (risk, limit)
        ).fetchall()
        return [dict(r) for r in rows]

    def get_alerts(self, hours: int = 24) -> list[dict]:
        since = (datetime.utcnow() - timedelta(hours=hours)).isoformat()
        rows  = self._conn().execute(
            "SELECT * FROM alerts WHERE alerted_at > ? ORDER BY id DESC",
            (since,)
        ).fetchall()
        return [dict(r) for r in rows]

    def get_stats(self) -> dict:
        c = self._conn()
        row = c.execute("""
            SELECT
                COUNT(*)                                AS total,
                SUM(CASE WHEN risk='critical' THEN 1 ELSE 0 END) AS critical,
                SUM(CASE WHEN risk='high'     THEN 1 ELSE 0 END) AS high,
                SUM(value_eth)                          AS eth_volume,
                COUNT(DISTINCT from_addr)               AS unique_wallets
            FROM transactions
        """).fetchone()
        return dict(row) if row else {}

    def search(self, wallet: str = "", func: str = "", risk: str = "", limit: int = 200) -> list[dict]:
        query  = "SELECT * FROM transactions WHERE 1=1"
        params = []
        if wallet:
            query += " AND (from_addr LIKE ? OR to_addr LIKE ?)"
            params += [f"%{wallet}%", f"%{wallet}%"]
        if func:
            query += " AND func_name LIKE ?"
            params.append(f"%{func}%")
        if risk:
            query += " AND risk=?"
            params.append(risk)
        query += " ORDER BY id DESC LIMIT ?"
        params.append(limit)
        rows = self._conn().execute(query, params).fetchall()
        return [dict(r) for r in rows]

    def total_count(self) -> int:
        return self._conn().execute("SELECT COUNT(*) FROM transactions").fetchone()[0]


# ═══════════════════════════════════════════════════════════════════════
# EXPORTERS
# ═══════════════════════════════════════════════════════════════════════
class MempoolExporter:
    """Export mempool data to CSV, JSON, and Excel."""

    def __init__(self, db: MempoolDB, export_dir: Path = EXPORT_DIR):
        self.db         = db
        self.export_dir = export_dir

    def _timestamp(self) -> str:
        return datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    def _rows_to_export(self, hours: int = 24, risk_filter: Optional[str] = None) -> list[dict]:
        since = (datetime.utcnow() - timedelta(hours=hours)).isoformat()
        c = self.db._conn()
        if risk_filter:
            rows = c.execute(
                "SELECT * FROM transactions WHERE seen_at > ? AND risk=? ORDER BY id DESC",
                (since, risk_filter)
            ).fetchall()
        else:
            rows = c.execute(
                "SELECT * FROM transactions WHERE seen_at > ? ORDER BY id DESC",
                (since,)
            ).fetchall()
        return [dict(r) for r in rows]

    # ── CSV ───────────────────────────────────────────────────────────
    def export_csv(self, hours: int = 24, risk_filter: Optional[str] = None) -> Path:
        rows = self._rows_to_export(hours, risk_filter)
        if not rows:
            print("[Export] No rows to export to CSV")
            return None

        suffix  = f"_{risk_filter}" if risk_filter else ""
        outpath = self.export_dir / f"mempool{suffix}_{self._timestamp()}.csv"

        with open(outpath, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)

        print(f"[Export] CSV → {outpath}  ({len(rows)} rows)")
        return outpath

    # ── JSON ──────────────────────────────────────────────────────────
    def export_json(self, hours: int = 24, risk_filter: Optional[str] = None) -> Path:
        rows = self._rows_to_export(hours, risk_filter)
        if not rows:
            print("[Export] No rows to export to JSON")
            return None

        suffix  = f"_{risk_filter}" if risk_filter else ""
        outpath = self.export_dir / f"mempool{suffix}_{self._timestamp()}.json"

        with open(outpath, "w") as f:
            json.dump({
                "exported_at": datetime.utcnow().isoformat(),
                "total_rows":  len(rows),
                "filter":      risk_filter or "all",
                "hours":       hours,
                "transactions": rows,
            }, f, indent=2, default=str)

        print(f"[Export] JSON → {outpath}  ({len(rows)} rows)")
        return outpath

    # ── Excel ─────────────────────────────────────────────────────────
    def export_excel(self, hours: int = 24) -> Optional[Path]:
        if not EXCEL_AVAILABLE:
            print("[Export] Install openpyxl and pandas for Excel export")
            return None

        rows = self._rows_to_export(hours)
        if not rows:
            print("[Export] No rows to export to Excel")
            return None

        outpath = self.export_dir / f"mempool_report_{self._timestamp()}.xlsx"
        df      = pd.DataFrame(rows)

        wb = openpyxl.Workbook()

        # ── Sheet 1: All Transactions ──────────────────────────────
        ws1 = wb.active
        ws1.title = "Transactions"
        _write_tx_sheet(ws1, df)

        # ── Sheet 2: Critical & High only ─────────────────────────
        df_alerts = df[df["risk"].isin(["critical", "high"])]
        if not df_alerts.empty:
            ws2 = wb.create_sheet("Alerts")
            _write_tx_sheet(ws2, df_alerts, highlight=True)

        # ── Sheet 3: Hourly stats ──────────────────────────────────
        ws3   = wb.create_sheet("Hourly Stats")
        since = (datetime.utcnow() - timedelta(hours=hours)).isoformat()
        stats = db._conn().execute(
            "SELECT * FROM stats_hourly WHERE hour > ? ORDER BY hour",
            (since[:13],)
        ).fetchall()
        if stats:
            ws3.append(["Hour (UTC)", "Total TX", "High Value", "Drain Sigs", "Critical", "ETH Volume"])
            for s in stats:
                ws3.append([s["hour"], s["total_tx"], s["high_value_tx"],
                             s["drain_sig_tx"], s["critical_tx"], round(s["total_eth"], 4)])
            _style_header(ws3)

        wb.save(outpath)
        print(f"[Export] Excel → {outpath}  ({len(rows)} rows)")
        return outpath

    # ── Export ALL formats at once ─────────────────────────────────
    def export_all(self, hours: int = 24) -> dict:
        return {
            "csv":   str(self.export_csv(hours)),
            "json":  str(self.export_json(hours)),
            "excel": str(self.export_excel(hours)) if EXCEL_AVAILABLE else "unavailable",
        }


# ── Excel helpers ──────────────────────────────────────────────────────
RISK_COLORS = {
    "critical": "FF2D2D",
    "high":     "EF4444",
    "medium":   "F59E0B",
    "low":      "6B7280",
}
HEADER_FILL = PatternFill("solid", fgColor="111827")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
THIN        = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

def _style_header(ws):
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border    = THIN
    ws.freeze_panes = "A2"

def _write_tx_sheet(ws, df: "pd.DataFrame", highlight: bool = False):
    COLS = ["hash", "from_addr", "to_addr", "value_eth", "gas_gwei",
            "func_name", "risk", "flags", "seen_at"]
    HEADERS = ["TX Hash", "From", "To", "Value (ETH)", "Gas (Gwei)",
               "Function", "Risk", "Flags", "Seen At (UTC)"]

    ws.append(HEADERS)
    _style_header(ws)

    for _, row in df[COLS].iterrows():
        ws.append(list(row))
        risk = str(row.get("risk", "low"))
        if highlight and risk in RISK_COLORS:
            color = RISK_COLORS[risk]
            for cell in ws[ws.max_row]:
                cell.fill   = PatternFill("solid", fgColor=color + "22")
                cell.border = THIN

    # Column widths
    for col, width in zip(ws.columns, [20, 16, 16, 12, 12, 14, 8, 30, 22]):
        ws.column_dimensions[col[0].column_letter].width = width


# ═══════════════════════════════════════════════════════════════════════
# AUTO-EXPORT SCHEDULER
# ═══════════════════════════════════════════════════════════════════════
import asyncio

class AutoExporter:
    """Exports all formats every N minutes in the background."""

    def __init__(self, exporter: MempoolExporter, interval_minutes: int = 15):
        self.exporter = exporter
        self.interval = interval_minutes * 60

    async def run(self):
        while True:
            await asyncio.sleep(self.interval)
            print(f"\n[AutoExport] Exporting last 24h data…")
            paths = self.exporter.export_all(hours=24)
            for fmt, path in paths.items():
                print(f"  {fmt.upper()} → {path}")


# ═══════════════════════════════════════════════════════════════════════
# SINGLETON (import and use directly)
# ═══════════════════════════════════════════════════════════════════════
db       = MempoolDB()
exporter = MempoolExporter(db)


# ── CLI usage ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    cmd = sys.argv[1] if len(sys.argv) > 1 else "stats"

    if cmd == "stats":
        s = db.get_stats()
        print(f"\n{'═'*40}")
        print(f"  Total TX stored : {s.get('total', 0):,}")
        print(f"  Critical        : {s.get('critical', 0):,}")
        print(f"  High risk       : {s.get('high', 0):,}")
        print(f"  ETH volume      : {s.get('eth_volume', 0):.4f} ETH")
        print(f"  Unique wallets  : {s.get('unique_wallets', 0):,}")
        print(f"{'═'*40}\n")

    elif cmd == "export":
        hours = int(sys.argv[2]) if len(sys.argv) > 2 else 24
        print(f"\n[Export] Exporting last {hours}h…")
        paths = exporter.export_all(hours=hours)
        print("\nFiles saved:")
        for fmt, p in paths.items():
            print(f"  {fmt.upper()}: {p}")

    elif cmd == "search":
        wallet = sys.argv[2] if len(sys.argv) > 2 else ""
        rows   = db.search(wallet=wallet)
        print(f"\nFound {len(rows)} transactions for wallet: {wallet}")
        for r in rows[:10]:
            print(f"  {r['hash'][:12]}… | {r['risk']:8} | {r['value_eth']:.4f} ETH | {r['seen_at']}")

    elif cmd == "alerts":
        alerts = db.get_alerts(hours=24)
        print(f"\nAlerts (last 24h): {len(alerts)}")
        for a in alerts:
            print(f"  [{a['risk'].upper()}] {a['tx_hash'][:12]}… | {a['value_eth']:.4f} ETH | {a['alerted_at']}")

    else:
        print("Usage: python mempool_storage.py [stats|export|search <wallet>|alerts]")
